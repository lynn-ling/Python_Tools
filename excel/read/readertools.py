import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

class excel_reader():
    def __init__(self):
        pass

    def file_reader(self, filePath):
        return file_reader(filePath)


class file_reader():
    def __init__(self, filePath):
        self.filePath = filePath
        self.workbook = openpyxl.load_workbook(filePath)
        pass

    def get_workbook(self):
        return self.workbook.get_sheet_names()

    def get_sheetnames(self):
        return self.workbook.get_sheet_names()


    def sum_by_rows(self, sheetName, row):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        sum = 0
        row_num = sheet[row]
        for cell in row_num:
            if cell.value != None:
                sum += float(cell.value)
        return sum



    def sum_by_column(self,sheetName, column):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        sum = 0
        column_num = sheet[column]
        for cell in column_num:
            if cell.value != None:
                sum += float(cell.value)
        return sum

    def avg_by_column(self,sheetName,column):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        sum = 0
        count = 0
        column_num = sheet[column]
        for cell in column_num:
            if cell.value != None:
                sum += float(cell.value)  
                count += 1      
        return sum/count 
    
    def avg_by_row(self,sheetName,row):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        sum = 0
        count = 0
        row_num = sheet[row]
        for cell in row_num:
            if cell.value != None:
                sum += float(cell.value)  
                count += 1      
        return sum/count    

    def max_by_column(self,sheetName,column):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        column_num = sheet[column]
        max = []
        for cell in column_num:
            if cell.value != None:
                max.append(cell.value)
    
        for i in range(len(max)-1):
            for j in range(len(max)-1-i):
                if max[j]>max[j+1]:
                    tmp = max[j]
                    max[j] = max[j+1]
                    max[j+1] = tmp
        return max[-1]
        
    
    def min_by_column(self,sheetName,column):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        column_num = sheet[column]
        min = []
        for cell in column_num:
            if cell.value != None:
                min.append(cell.value)
    
        for i in range(len(min)-1):
            for j in range(len(min)-1-i):
                if min[j]<min[j+1]:
                    tmp = min[j]
                    min[j] = min[j+1]
                    min[j+1] = tmp
        return min[-1]



