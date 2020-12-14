import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

class excel_writer():
    def __init__(self):
        pass

    def file_writer(self, filePath):
        return file_writer(filePath)


class file_writer():
    def __init__(self, filePath):
        self.filePath = filePath
        self.workbook = openpyxl.load_workbook(filePath)
        pass

    def save(self):
        self.workbook.save(self.filePath)
    
    def create_sheet(self, name):
        newsheet = self.workbook.create_sheet(title=name)
        return newsheet

    def add_data(self, sheet, data):
        for rownum in range(0, len(data)):
            for column in range(0,len(data[rownum]):
                pass