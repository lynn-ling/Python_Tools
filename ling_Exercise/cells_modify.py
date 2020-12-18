import openpyxl
from openpyxl.styles import Font,PatternFill,Border,Side,Color,Alignment

work_sheet = openpyxl.load_workbook("C://Users//baozipu//Desktop//sum.xlsx")
sheet = work_sheet.get_sheet_by_name("合计")

fontObj=Font(name=u'华文新魏',size=20, color='00ff00', bold=True, italic=True, strike=True , underline="double")
# name: 字体的名字，如果为中文必须加<u>；  size：字体大小；  color：颜色
# bole：加粗 ；  italic：斜体；  strike：加删除线;  underline：加下划线
# 颜色对照可以参考网址：https://m.wang1314.com/doc/webapp/topic/21084865.html

def font_modify(start,end,font):
    for tuples in sheet[start+":"+end]:
        for cells in tuples:
            cells.font = font

#font_modify("A2","B10",fontObj)


patternObj=PatternFill(fill_type="solid",fgColor="00ff00")

#fill_type 有如下的方式 一般纯色填充使用 solid， 其他样式自行尝试
#{'lightGrid', 'gray0625', 'lightTrellis', 'lightDown', 'lightVertical', 
#'darkTrellis', 'darkHorizontal', 'darkVertical', 'darkGrid', 'darkGray',
#'solid', 'darkUp', 'lightGray', 'mediumGray', 'darkDown', 'lightHorizontal', 'lightUp', 'gray125'}

def PatternFill_modify(start,end,patternFill):
    for tuples in sheet[start+":"+end]:
        for cells in tuples:
            cells.fill = patternFill

#PatternFill_modify("A3","A3",patternObj)

borderObj = Border(left=Side(border_style='thin',color='00ff00'),right=Side(border_style='dashDot',color='00ff00'),
            top=Side(border_style='double',color='00ff00'),bottom=Side(border_style='hair',color='00ff00'))

#border_style样式很多：‘dashDot’,‘dashDotDot’,‘dashed’,‘dotted’,‘double’,‘hair’,‘medium’,
#‘mediumDashDot’,‘mediumDashDotDot’,‘mediumDashed’,‘slantDashDot’,‘thick’,‘thin’

def border_modify(start,end,border):
        for tuples in sheet[start+":"+end]:
            for cells in tuples:
                cells.border = border

#border_modify("A3","B6",borderObj)

alignmentObj = Alignment(horizontal='left',vertical='top',wrap_text=True)
#horizontal代表水平方向，可以左对齐left，还有居中center和右对齐right，分散对齐distributed，跨列居中centerContinuous，两端对齐justify，填充fill，常规general
#vertical代表垂直方向，可以居中center，还可以靠上top，靠下bottom，两端对齐justify，分散对齐distributed
#自动换行：wrap_text，这是个布尔类型的参数，这个参数还可以写作wrapText
 
def alignment_change(start,end,alignment):       
        for tuples in sheet[start+":"+end]:
            for cells in tuples:
                cells.alignment = alignment

#alignment_change("A3","B6",alignmentObj)

work_sheet.save("C://Users//baozipu//Desktop//sum.xlsx")


