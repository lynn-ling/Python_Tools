import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


work_book = openpyxl.load_workbook("C://Users//baozipu//Desktop//sum.xlsx")
sheet = work_book.get_sheet_by_name("合计")

column_max = sheet.max_column
row_max = sheet.max_row


for i in range (1,column_max+1):
    column_num = get_column_letter(i)
    sum = 0
    for j in range(1,row_max+1):
        if sheet[column_num+str(j)].value != None:
            sum+=sheet[column_num+str(j)].value
    sheet[column_num+str(row_max+1)] = sum
work_book.save("C://Users//baozipu//Desktop//sum.xlsx")
    
