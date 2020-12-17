import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

#代码1）
work_book = openpyxl.load_workbook("C://Users//baozipu//Desktop//sum.xlsx",data_only=True)
sheet = work_book.get_sheet_by_name("合计")

column_max = sheet.max_column
row_max = sheet.max_row

for i in range (1,column_max+1):
    column_num = get_column_letter(i)
    sheet[column_num+str(row_max+1)] = '=SUM({}1:{}{})'.format(column_num,column_num,row_max)
    #必须要用{}.format的形式向公式里传参数，不然最后合计的格子里会显示 =sum(column_range)，结果为?NAME错误
    #column_range = column_num+str(1)+":"+column_num+str(row_max+1)
    #sheet[column_num+str(row_max+1)] = '=sum(column_range)'
work_book.save("C://Users//baozipu//Desktop//sum.xlsx")


#代码2）
work_book = openpyxl.load_workbook("C://Users//baozipu//Desktop//sum.xlsx",data_only=True)
sheet = work_book.get_sheet_by_name("合计")
print(sheet['A24'].value) 
#将代码1）2）一次性执行取出来的值为None。这时需要先执行代码1），然后手动打开Excel表，保存
#再执行代码2）才能取到正确的值



