import openpyxl
from openpyxl.chart import BarChart, Series, Reference


work_sheet = openpyxl.load_workbook("C://Users//baozipu//Desktop//sum.xlsx")
sheet = work_sheet.get_sheet_by_name("chart")

#type：为col表示垂直，为bar表示水平
#style：样式，数值范围为1~48；   title：标题
#y_axis.title：纵坐标标题；x_axis.title：横坐标标题
#data为数据引用范围
#cats为类别对象

max_column = sheet.max_column
max_row = sheet.max_row

def chart_generate(types,style,title,y_title,x_title):  
    chart1 = BarChart()
    chart1.type = types
    chart1.style = style
    chart1.title = title
    chart1.y_axis.title = y_title
    chart1.x_axis.title = x_title

    data = Reference(sheet, min_col=2, min_row=1, max_row=max_row, max_col=max_column)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    sheet.add_chart(chart1, "A10")
    work_sheet.save("C://Users//baozipu//Desktop//sum.xlsx")

chart_generate("col",20,"bar chart","QTY","fruits")