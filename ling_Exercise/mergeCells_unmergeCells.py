import openpyxl

work_book = openpyxl.load_workbook("C://Users//baozipu//Desktop//sum.xlsx")
sheet = work_book.get_sheet_by_name("合计")

def merge_cells(start,end):
    sheet.merge_cells(start+":"+end)
    work_book.save("C://Users//baozipu//Desktop//sum.xlsx")

merge_cells("A2","B5")

def unmerge_cell(start,end):
    sheet.unmerge_cells(start+":"+end)
    work_book.save("C://Users//baozipu//Desktop//sum.xlsx")

unmerge_cell("H10","I11")


